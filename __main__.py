import sys
import random
import string
import openpyxl
import os

MAX_CODE_LEN = 12
EXCEL_FILE_NAME = "code.xlsx"

def GenCode(count):
    old_list = []
    sheet = wb["code"]
    if sheet:

    for i in range(count):
        ran_str = ''.join(random.sample(
            string.ascii_letters + string.digits, MAX_CODE_LEN))
        if ran_str not in old_list:
            old_list.append(ran_str)
    WriteToExcel(old_list)


def LoadExcel():
    global wb
    if os.path.exists(EXCEL_FILE_NAME):
        print("load excel")
        wb = openpyxl.load_workbook(EXCEL_FILE_NAME)
    else:
        print("create execel")
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("code")
        print(wb.sheetnames)


def WriteToExcel(list):
    sheet = wb.get_sheet_by_name("code")
    for i in range(len(list)):
        key = "A"+i
        sheet[key] = list[i]
    wb.save(EXCEL_FILE_NAME)



if __name__ == "__main__":
    args = sys.argv
    LoadExcel()
    GenCode(int(args[1]))
